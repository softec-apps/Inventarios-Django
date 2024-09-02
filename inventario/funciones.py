#----------------------------FUNCIONES DE AYUDA Y COMPLEMENTO--------------------------------------------------

from .models import Producto, Opciones, Kardex, Proveedor, Pedido, Categoria, Cliente, Usuario, Factura, Descuento
from decimal import Decimal
import csv
from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime


def obtenerIdProducto(descripcion):
    id_producto = Producto.objects.get(descripcion=descripcion)
    resultado = id_producto.id

    return resultado

def productoTieneIva(idProducto):
    iva = Producto.objects.get(id=idProducto)
    resultado = iva.tiene_iva

    return resultado

def sacarIva(elemento):
    iva = Opciones.objects.get(id=1)
    ivaSacado =  iva.valor_iva/100
    resultado = elemento + (elemento * Decimal(ivaSacado))
    return resultado

def ivaActual(modo):
    if modo == 'valor':
        iva = Opciones.objects.get(id=1)
        return iva.valor_iva

    elif modo == 'objeto':
        iva = Opciones.objects.get(id=1)
        return iva

def obtenerProducto(idProducto):
    producto = Producto.objects.get(id=idProducto)
    return producto


def complementarContexto(contexto,datos):
    contexto['usuario'] = datos.username
    contexto['id_usuario'] = datos.id
    contexto['nombre'] = datos.first_name
    contexto['apellido'] = datos.last_name
    contexto['correo'] = datos.email

    return contexto

def usuarioExiste(Usuario,buscar,valor):
    if buscar == 'username':
        try:
            Usuario.objects.get(username=valor)
            return True
        except Usuario.DoesNotExist:
            return False

    elif buscar == 'email':
        try:
            Usuario.objects.get(email=valor)
            return True
        except Usuario.DoesNotExist:
            return False

def manejarArchivo(archivo,ruta):
    with open(ruta, 'wb+') as destino:
        for chunk in archivo.chunks():
            destino.write(chunk)


def render_to_pdf(template_src, context_dict={}):
    import io
    from xhtml2pdf import pisa
    from django.template.loader import get_template
    from django.http import HttpResponse

    template = get_template(template_src)
    html  = template.render(context_dict)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None

#--------------------------------------------------------------------------------------------------------------                 

#----------------------------FUNCIONES PARA LA EXPORTACION DE ARCHIVOS SCV, EXEL--------------------------------------------------

def exportar_productos_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="productos.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Descripcion', 'Precio', 'Disponible', 'Medida', 'Categoria'])

    productos = Producto.objects.all().values_list('id', 'descripcion', 'precio', 'disponible', 'medida', 'categoria__nombre')
    for producto in productos:
        writer.writerow(producto)

    return response

def exportar_productos_excel(request):
    productos = Producto.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'

    columns = ['ID', 'Descripción', 'Precio', 'Disponible', 'Medida', 'Categoría']
    ws.append(columns)

    for producto in productos:
        ws.append([
            producto.id,
            producto.descripcion,
            producto.precio,
            producto.disponible,
            producto.get_medida_display(),
            producto.categoria.nombre
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'

    wb.save(response)
    return response


def exportar_kardex_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="kardex.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Producto', 'Fecha', 'Tipo Movimiento', 'Cantidad', 'Valor Unitario', 'Valor Total', 'Saldo Cantidad', 'Saldo Valor Total', 'Detalle'])
    
    kardex_entries = Kardex.objects.all().select_related('producto')
    for entry in kardex_entries:
        writer.writerow([
            entry.id, entry.producto.descripcion, entry.fecha, entry.tipo_movimiento,
            entry.cantidad, entry.valor_unitario, entry.valor_total,
            entry.saldo_cantidad, entry.saldo_valor_total, entry.detalle
        ])
    
    return response

def exportar_kardex_excel(request):
    kardex_entries = Kardex.objects.all().select_related('producto')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Kardex"
    
    ws.append(['ID', 'Producto', 'Fecha', 'Tipo Movimiento', 'Cantidad', 'Valor Unitario', 'Valor Total', 'Saldo Cantidad', 'Saldo Valor Total', 'Detalle'])
    
    for entry in kardex_entries:
        # Formatear la fecha
        fecha_formateada = entry.fecha.strftime('%d-%m-%Y')
        ws.append([
            entry.id, entry.producto.descripcion, fecha_formateada, entry.tipo_movimiento,
            entry.cantidad, entry.valor_unitario, entry.valor_total,
            entry.saldo_cantidad, entry.saldo_valor_total, entry.detalle
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=kardex.xlsx'
    wb.save(response)
    return response

def exportar_proveedores_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="proveedores.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Tipo Cédula', 'Cédula', 'Nombre', 'Apellido', 'Ciudad', 'Dirección', 'Teléfono', 'Teléfono 2', 'Correo', 'Correo 2', 'RUC'])
    
    proveedores = Proveedor.objects.all()
    for proveedor in proveedores:
        writer.writerow([
            proveedor.id, proveedor.get_tipoCedula_display(), proveedor.cedula,
            proveedor.nombre, proveedor.apellido, proveedor.ciudad, proveedor.direccion,
            proveedor.telefono, proveedor.telefono2, proveedor.correo, proveedor.correo2,
            proveedor.ruc
        ])
    
    return response

def exportar_proveedores_excel(request):
    proveedores = Proveedor.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    ws.append(['ID', 'Tipo Cédula', 'Cédula', 'Nombre', 'Apellido', 'Ciudad', 'Dirección', 'Teléfono', 'Teléfono 2', 'Correo', 'Correo 2', 'RUC'])
    
    for proveedor in proveedores:
        ws.append([
            proveedor.id, proveedor.get_tipoCedula_display(), proveedor.cedula,
            proveedor.nombre, proveedor.apellido, proveedor.ciudad, proveedor.direccion,
            proveedor.telefono, proveedor.telefono2, proveedor.correo, proveedor.correo2,
            proveedor.ruc
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'
    wb.save(response)
    return response

def exportar_pedidos_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="pedidos.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Proveedor', 'Fecha', 'Sub Monto', 'Monto General', 'Recibido'])
    
    pedidos = Pedido.objects.all().select_related('proveedor')
    for pedido in pedidos:
        writer.writerow([
            pedido.id, f"{pedido.proveedor.nombre} {pedido.proveedor.apellido}",
            pedido.fecha, pedido.sub_monto, pedido.monto_general, 'Si' if pedido.presente else 'No'
        ])
    
    return response

def exportar_pedidos_excel(request):
    pedidos = Pedido.objects.all().select_related('proveedor')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    
    ws.append(['ID', 'Proveedor', 'Fecha', 'Sub Monto', 'Monto General', 'Recibido'])
    
    for pedido in pedidos:
        fecha_formateada = pedido.fecha.strftime('%d-%m-%Y')
        ws.append([
            pedido.id, f"{pedido.proveedor.nombre} {pedido.proveedor.apellido}",
            fecha_formateada, pedido.sub_monto, pedido.monto_general, 'Si' if pedido.presente else 'No'
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pedidos.xlsx'
    wb.save(response)
    return response

def exportar_categorias_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="categorias.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Nombre'])
    
    categorias = Categoria.objects.all()
    for categoria in categorias:
        writer.writerow([categoria.id, categoria.nombre])
    
    return response

def exportar_categorias_excel(request):
    categorias = Categoria.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorias"
    
    ws.append(['ID', 'Nombre'])
    
    for categoria in categorias:
        ws.append([categoria.id, categoria.nombre])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=categorias.xlsx'
    wb.save(response)
    return response

def exportar_clientes_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="clientes.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Tipo Cédula', 'Cédula', 'Nombre', 'Apellido', 'Ciudad', 'Dirección', 'Teléfono', 'Teléfono 2', 'Correo', 'Correo 2', 'Género', 'Nacimiento'])
    
    clientes = Cliente.objects.all()
    for cliente in clientes:
        writer.writerow([
            cliente.id, cliente.get_tipoCedula_display(), cliente.cedula,
            cliente.nombre, cliente.apellido, cliente.ciudad, cliente.direccion,
            cliente.telefono, cliente.telefono2, cliente.correo, cliente.correo2,
            cliente.get_genero_display(), cliente.nacimiento
        ])
    
    return response

def exportar_clientes_excel(request):
    clientes = Cliente.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    ws.append(['ID', 'Tipo Cédula', 'Cédula', 'Nombre', 'Apellido', 'Ciudad', 'Dirección', 'Teléfono', 'Teléfono 2', 'Correo', 'Correo 2', 'Género', 'Nacimiento'])
    
    for cliente in clientes:
        ws.append([
            cliente.id, cliente.get_tipoCedula_display(), cliente.cedula,
            cliente.nombre, cliente.apellido, cliente.ciudad, cliente.direccion,
            cliente.telefono, cliente.telefono2, cliente.correo, cliente.correo2,
            cliente.get_genero_display(), cliente.nacimiento
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=clientes.xlsx'
    wb.save(response)
    return response

def exportar_descuentos_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="descuentos.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Nombre Descuento', 'Valor'])
    
    descuentos = Descuento.objects.all()
    for descuento in descuentos:
        writer.writerow([
            descuento.id,
            descuento.nombre,
            descuento.valor
        ])
    
    return response

def exportar_descuentos_excel(request):
    descuentos = Descuento.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Descuentos"
    
    ws.append(['ID', 'Nombre Descuento', 'Valor'])
    
    for descuento in descuentos:
        ws.append([
            descuento.id,
            descuento.nombre,
            descuento.valor
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=descuentos.xlsx'
    wb.save(response)
    return response

def exportar_facturas_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition']='attachment; filename="ventas.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Cliente', 'Fecha','Sub Monto','Monto General'])

    facturas = Factura.objects.all().select_related('cliente')
    for factura in facturas:
        writer.writerow([
            factura.id,
            f"{factura.cliente.nombre} {factura.cliente.apellido}",
            factura.fecha,
            factura.sub_monto,
            factura.monto_general
        ])

    return response

def exportar_facturas_excel(request):
    facturas = Factura.objects.all().select_related('cliente')

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    columns = ['ID', 'Cliente', 'Fecha', 'Sub Monto', 'Monto general']
    ws.append(columns)

    for factura in facturas:
        fecha_formateada = factura.fecha.strftime('%d-%m-%Y')
        ws.append([
            factura.id,
            f"{factura.cliente.nombre} {factura.cliente.apellido}",
            fecha_formateada,
            factura.sub_monto,
            factura.monto_general
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'

    wb.save(response)
    return response

def exportar_usuarios_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="usuarios.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Username', 'Email', 'Nombre', 'Apellido', 'Nivel', 'Es Superusuario'])
    
    usuarios = Usuario.objects.all()
    for usuario in usuarios:
        writer.writerow([
            usuario.id, usuario.username, usuario.email,
            usuario.first_name, usuario.last_name, usuario.nivel,
            usuario.is_superuser
        ])
    
    return response

def exportar_usuarios_excel(request):
    usuarios = Usuario.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    ws.append(['ID', 'Username', 'Email', 'Nombre', 'Apellido', 'Nivel', 'Es Superusuario'])
    
    for usuario in usuarios:
        ws.append([
            usuario.id, usuario.username, usuario.email,
            usuario.first_name, usuario.last_name, usuario.nivel,
            usuario.is_superuser
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
    wb.save(response)
    return response

